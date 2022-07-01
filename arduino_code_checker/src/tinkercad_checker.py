import datetime
import configparser
import argparse
import sys
import os
from playwright.sync_api import Playwright, sync_playwright, expect
from arduino_code_checker.src.code_check import batch_compare
from arduino_code_checker.src.download_code import run
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl

if __name__ == "__main__":
    start = datetime.datetime.now()
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument(
        "--config_file", help="Configuration file", type=str, default="config.ini"
    )
    arg_parser.add_argument(
        "--solutions_folder", help="Solutions folder", type=str, default="solutions"
    )
    arg_parser.add_argument(
        "--download_folder", help="Download folder", type=str, default=""
    )

    args = arg_parser.parse_args().__dict__

    parser = configparser.ConfigParser()
    parser.read(args.get("config_file"))

    try:
        student_list = parser["tinkercad"]["students"].replace(" ", "").split(",")
        if len(student_list) == 1 and len(student_list[0]) == 0:
            student_list = []
    except KeyError:
        student_list = []

    try:
        headless = parser["tinkercad"].getboolean("headless")
    except KeyError:
        headless = True

    assignment_list = parser["tinkercad"]["questions"].replace(" ", "").split(",")

    if len(args.get("download_folder")) == 0:
        with sync_playwright() as playwright:
            download_folder = run(
                playwright,
                classroom_url=parser["tinkercad"]["classroom"],
                user=parser["tinkercad"]["user"],
                password=parser["tinkercad"]["password"],
                student_list=student_list,
                assigment_list=assignment_list,
                headless=headless,
            )
    else:
        download_folder = args.get("download_folder")

    result = batch_compare(
        args.get("solutions_folder"), download_folder, assignment_list
    )
    result.drop(result.columns[1], axis=1, inplace=True)
    result.sort_values(by=[result.columns[0], result.columns[1]], inplace=True)

    codigo_col_index = list(result.columns).index("codigo") + 1
    circuito_col_index = list(result.columns).index("circuito") + 1
    error_col_index = list(result.columns).index("erro_codigo") + 1

    wb = openpyxl.Workbook()
    ws = wb.active

    for r in dataframe_to_rows(result, index=False, header=True):
        ws.append(r)

    from openpyxl.styles import Font

    fontStyle = Font(size="13")

    for title_index in range(1, 20):
        ws.cell(row=1, column=title_index).font = fontStyle

    for row_index in range(2, len(result) + 2):
        cod_cell = ws.cell(row=row_index, column=codigo_col_index)
        circ_cell = ws.cell(row=row_index, column=circuito_col_index)
        ws.cell(row=row_index, column=error_col_index).alignment = Alignment(
            wrap_text=True
        )
        ws.cell(row=row_index, column=circuito_col_index + 1).value = " "

        if cod_cell.value is not None and ".ino" in cod_cell.value:
            cod_cell.hyperlink = f"file:///{os.path.join(os.getcwd(), cod_cell.value)}"
            cod_cell.style = "Hyperlink"

        if circ_cell.value is not None and ".brd" in circ_cell.value:
            circ_cell.hyperlink = (
                f"file:///{os.path.join(os.getcwd(), circ_cell.value)}"
            )
            circ_cell.style = "Hyperlink"

    ws.column_dimensions["A"].width = 30
    for col_index in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_index].width = 25

    tab = Table(displayName="Table1", ref=f"A1:I{len(result) + 2}")

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style

    ws.add_table(tab)

    wb.save(f"evaluation_{datetime.datetime.now().strftime('%Y-%m-%d-%H_%M_%S')}.xlsx")
    end = datetime.datetime.now()

    print(f"Time elapsed: {end - start}s")

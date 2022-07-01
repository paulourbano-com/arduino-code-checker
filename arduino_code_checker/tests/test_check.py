import os
import pandas as pd
from arduino_code_checker.src.code_check import (
    compare_solution_assignment,
    compare_set_inner_func_calls,
    batch_compare,
)
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl


def test_check_01():
    result = compare_solution_assignment(
        os.path.join(os.path.dirname(__file__), "solution_01", "solution.ino"),
        os.path.join(os.path.dirname(__file__), "assigments_01", "assigment.ino"),
    )

    assert sum(result.values()) / len(result) == 100.0


def test_non_existant_file():
    result = compare_solution_assignment(
        os.path.join(os.path.dirname(__file__), "solution_01", "solution.ino"),
        os.path.join(os.path.dirname(__file__), "assigments_01", "not_there.ino"),
    )

    assert sum(result.values()) / len(result) == 0.0


def test_parse_error():
    result = compare_solution_assignment(
        os.path.join(os.path.dirname(__file__), "solution_01", "solution.ino"),
        os.path.join(
            os.path.dirname(__file__), "assigments_01", "assigment_syntax_error.ino"
        ),
    )

    assert sum(result.values()) / len(result) == 0.0


def test_missing_definition():
    result = compare_solution_assignment(
        os.path.join(os.path.dirname(__file__), "solution_01", "solution.ino"),
        os.path.join(
            os.path.dirname(__file__), "assigments_01", "assigment_missing_def.ino"
        ),
    )

    assert sum(result.values()) / len(result) == 50.0


def test_compare_setups():
    percent_match = compare_set_inner_func_calls(
        "setup",
        os.path.join(os.path.dirname(__file__), "solution_01", "solution.ino"),
        os.path.join(
            os.path.dirname(__file__),
            "assigments_01",
            "assigment_multiple_setup_calls.ino",
        ),
    )

    assert percent_match == 100.0


def test_compare_folders():
    result = batch_compare(
        "solutions",
        "tinkercad_downloads_2022-06-30-21_15_06",
        ["E1T1", "E1T2", "E2T1", "E2T2", "E3T1"],
    )

    result.sort_values(by=result.columns[0], inplace=True)

    codigo_col_index = list(result.columns).index("codigo") + 1
    circuito_col_index = list(result.columns).index("circuito") + 1

    wb = openpyxl.Workbook()
    ws = wb.active

    for r in dataframe_to_rows(result, index=False, header=True):
        ws.append(r)

    for row_index in range(3, len(result) + 2):
        cod_cell = ws.cell(row=row_index, column=codigo_col_index)
        circ_cell = ws.cell(row=row_index, column=circuito_col_index)

        if cod_cell.value is not None and ".ino" in cod_cell.value:
            cod_cell.hyperlink = f"file:///{os.path.join(os.getcwd(), cod_cell.value)}"
            cod_cell.style = "Hyperlink"

        if circ_cell.value is not None and ".brd" in circ_cell.value:
            circ_cell.hyperlink = (
                f"file:///{os.path.join(os.getcwd(), circ_cell.value)}"
            )
            circ_cell.style = "Hyperlink"

        # print(ws.cell(row=row_index, column=codigo_col_index).value)
        # print(ws.cell(row=row_index, column=circuito_col_index).value)
        # print()

    ws.column_dimensions["A"].width = 30
    for col_index in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_index].width = 20

    wb.save("test_result.xlsx")

    assert isinstance(result, pd.DataFrame)
    assert len(result) > 0
